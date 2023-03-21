import openpyxl as xl
import os
"""to communicate with os """


class Searchdrives():
    """this module helps to find all drives in pc
    67 to 91 for ascii values
    char() is to convert the ascii value to character!"""
    def __init__(self):
        self.drives = []
        self.workbook = xl.load_workbook("C://testdata//Testdrives.xlsx")

    def searchmydrives(self) -> list:
        for x in range(67, 91):
            if os.path.exists(chr(x) + ":"):
                self.drives.append(chr(x))

        return self.drives


dr = Searchdrives()
data = str(dr.searchmydrives())
print(data)
sheet = dr.workbook.active
sheet.cell(row=1, column=1).value = data
sheet.cell(row=2, column=1).value = data
sheet.cell(row=3, column=1).value = data
dr.workbook.save("C://testdata//Testdrives.xlsx")
dr.workbook.close()

