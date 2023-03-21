import mysql.connector
import time
import openpyxl as xl
import logging
logging.basicConfig(filename='..//logger//capstone.log', level=logging.DEBUG, format='%(asctime)s-%(name)s-%(levelname)s-%(message)s')

from SearchFilesinDrives.Searchfiles import searchfilesdrives
from SearchinDB.DBConnection import DatabaseConnection
from SearchinDB.searchfilepath import SearchFiles
from capstonexception.Mysqlexception import MYSQLError
from SearchinDB.InsertData import InsertinDB


def userdata():
    dir = input("enter the drive like c:// d:// \n")
    filename = input("enter the filename with extension like demo.txt \n")
    logging.info(f"drive name{dir} file name  {filename}")
    dbobj = SearchFiles()
    logging.info(f"class used {SearchFiles.__name__}")
    wb = xl.load_workbook("C://testdata//Testfiles.xlsx")
    ws = wb.active

    try:
        dbobj.connect("localhost","root","Navya@123","myhcl",3306)
        logging.info("myhcl database is connected")
        result = dbobj.search(filename)
    except mysql.connector.Error as err:
        logging.exception(err, exc_info=True)
        raise MYSQLError(f'{err.msg}', err.errno)
    finally:
        dbobj.connect.close()

    if len(result) == 0:
        print("Not Found in Database")
        print("Now searching in Drives...!")
        logging.info("not found in database")
        logging.info("not found in Drives")
        start_time=time.time()
        obj=searchfilesdrives()
        logging.info(f' for searching in drive  {searchfilesdrives.__name__} is used')
        files=obj.searchfiles(dir,filename)
        ws.cell(row=1, column=1).value = str(files)
        # sheet.cell(row=2, column=1).value = data
        # sheet.cell(row=3, column=1).value = data
        wb.save("C://testdata//Testfiles.xlsx")
        wb.close()
        inserobj=InsertinDB()
        inserobj.insert(files)
        print(files)
        obj.start()
        end_time=time.time()
        logging.info(f'time taken{end_time-start_time}')
        logging.info("ending")
        print(end_time-start_time)
    else:
        result="found in database!"
        print(result)


userdata()
